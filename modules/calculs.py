"""
Module Calculs — Stats, absences, leaderboard
Toute la logique métier CrossFit CLC
"""

import openpyxl
import logging
from datetime import datetime, timedelta
from collections import defaultdict

log = logging.getLogger(__name__)

# Activités exclues du leaderboard (accès libre)
ACTIVITES_EXCLUES = {"Open Gym", "Zone Haltéro", "GYM"}

# Seuils d'absence en jours
SEUIL_7J  = 7
SEUIL_14J = 14
SEUIL_21J = 21


def calculer_stats(fichiers: dict) -> dict:
    """
    Calcule les statistiques complètes à partir des 4 fichiers.
    Retourne un dict avec profil complet de chaque membre.
    """
    membres   = _charger_membres(fichiers["utilisateurs"])
    abonnes   = _charger_abonnements(fichiers["abonnements"])
    cartes    = _charger_cartes(fichiers["cartes"])
    presences = _charger_presences(fichiers["presences"])

    today = datetime.now()
    mois_actuel = f"{today.year}-{today.month:02d}"

    stats = {}

    for idd, membre in membres.items():
        pres_membre = presences.get(idd, {})
        abo = abonnes.get(idd)
        carte = cartes.get(idd)

        # Séances par mois (cours coachés uniquement)
        seances_par_mois = {}
        for mois, nb in pres_membre.items():
            seances_par_mois[mois] = nb

        # Dernière présence
        derniere = _derniere_presence(fichiers["presences"], idd)

        # Jours d'absence
        jours_absence = (today - derniere).days if derniere else 9999

        # Statut payant
        est_payant = bool(abo) or bool(carte and carte.get("creneaux", 0) > 0)

        stats[idd] = {
            "identifiant":     idd,
            "prenom":          membre.get("prenom", ""),
            "nom":             membre.get("nom", ""),
            "email":           membre.get("email", ""),
            "telephone":       membre.get("telephone", ""),
            "abonnement":      abo.get("type") if abo else None,
            "abo_expiration":  abo.get("expiration") if abo else None,
            "carte_creneaux":  carte.get("creneaux") if carte else None,
            "carte_expiration":carte.get("expiration") if carte else None,
            "est_payant":      est_payant,
            "seances_par_mois":seances_par_mois,
            "seances_mois_actuel": seances_par_mois.get(mois_actuel, 0),
            "derniere_presence":   derniere.strftime("%d/%m/%Y") if derniere else None,
            "jours_absence":       jours_absence,
        }

    log.info(f"Stats calculées pour {len(stats)} membres")
    return stats


def detecter_absences(fichiers: dict) -> dict:
    """
    Détecte les membres absents selon les seuils,
    en ne gardant que les membres avec abonnement actif.
    """
    stats = calculer_stats(fichiers)
    absences = {"7j": [], "14j": [], "21j": []}

    for idd, s in stats.items():
        if not s["est_payant"]:
            continue
        j = s["jours_absence"]
        if SEUIL_7J <= j < SEUIL_14J:
            absences["7j"].append(s)
        elif SEUIL_14J <= j < SEUIL_21J:
            absences["14j"].append(s)
        elif j >= SEUIL_21J:
            absences["21j"].append(s)

    log.info(f"Absences — 7j:{len(absences['7j'])} 14j:{len(absences['14j'])} 21j+:{len(absences['21j'])}")
    return absences


def generer_leaderboard(fichiers: dict, top_n: int = 15) -> list:
    """
    Génère le leaderboard du mois en cours.
    Exclut les accès libres (Open Gym, Zone Haltéro, GYM).
    Retourne une liste triée des top membres.
    """
    today = datetime.now()
    mois_actuel = f"{today.year}-{today.month:02d}"
    mois_precedent = _mois_precedent(today)

    # Charger présences filtrées
    seances_mois  = defaultdict(int)
    seances_prec  = defaultdict(int)
    membres_info  = {}

    wb_u = openpyxl.load_workbook(fichiers["utilisateurs"])
    ws_u = wb_u.active
    for i, row in enumerate(ws_u.iter_rows(values_only=True)):
        if i < 2: continue
        idd = row[9]
        if idd:
            membres_info[idd] = {
                "prenom": row[3] or "",
                "nom":    row[4] or "",
                "initiale": (row[4] or " ")[0].upper() + "."
            }

    wb_p = openpyxl.load_workbook(fichiers["presences"])
    ws_p = wb_p.active

    # Lire headers presences
    pres_headers = {}
    for i, row in enumerate(ws_p.iter_rows(values_only=True)):
        if i == 0: continue
        if i == 1:
            for j, col in enumerate(row):
                if col:
                    pres_headers[str(col).strip()] = j
            break

    p_idd      = pres_headers.get("Identifiant", 1)
    p_activite = pres_headers.get("Activité / Catégorie", 4)
    p_date     = pres_headers.get("Date", 5)
    p_presence = pres_headers.get("Présence", 13)

    for i, row in enumerate(ws_p.iter_rows(values_only=True)):
        if i < 2: continue
        if len(row) <= max(p_idd, p_activite, p_date, p_presence): continue

        idd      = row[p_idd]
        activite = row[p_activite] if len(row) > p_activite else None
        date_str = row[p_date] if len(row) > p_date else None
        presence = row[p_presence] if len(row) > p_presence else None

        if presence != "Oui" or not idd or not date_str:
            continue
        if activite in ACTIVITES_EXCLUES:
            continue
        if idd in ("nubguest1", "nubguest2", "nubguest3", "nubguest4"):
            continue

        try:
            date = datetime.strptime(str(date_str), "%d-%m-%Y")
            mois = f"{date.year}-{date.month:02d}"
            if mois == mois_actuel:
                seances_mois[idd] += 1
            elif mois == mois_precedent:
                seances_prec[idd] += 1
        except:
            pass

    # Construire le leaderboard
    leaderboard = []
    for rang, (idd, nb) in enumerate(
        sorted(seances_mois.items(), key=lambda x: -x[1])[:top_n], 1
    ):
        info = membres_info.get(idd, {})
        nb_prec = seances_prec.get(idd, 0)
        leaderboard.append({
            "rang":         rang,
            "identifiant":  idd,
            "prenom":       info.get("prenom", idd),
            "nom_initial":  info.get("initiale", ""),
            "seances":      nb,
            "seances_prec": nb_prec,
            "progression":  nb - nb_prec,
        })

    return leaderboard


# ── Fonctions internes ─────────────────────────────────────────────────────────

def _charger_membres(fichier: str) -> dict:
    wb = openpyxl.load_workbook(fichier)
    ws = wb.active
    membres = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue
        idd = row[9]
        if idd:
            membres[idd] = {
                "prenom":    row[3] or "",
                "nom":       row[4] or "",
                "email":     row[6] or "",
                "telephone": row[2] or "",
                "statut":    row[8] or "",
            }
    return membres


def _charger_abonnements(fichier: str) -> dict:
    wb = openpyxl.load_workbook(fichier)
    ws = wb.active
    abonnes = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue
        idd    = row[1]
        statut = row[7]
        if idd and statut == "Active":
            abonnes[idd] = {
                "type":       row[2],
                "profil":     row[3],
                "expiration": row[10],
            }
    return abonnes


def _charger_cartes(fichier: str) -> dict:
    wb = openpyxl.load_workbook(fichier)
    ws = wb.active
    cartes = {}

    # Lire les headers pour trouver les bons index
    headers = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if i == 1:
            for j, col in enumerate(row):
                if col:
                    headers[str(col).strip()] = j
            break

    idx_idd      = headers.get("Identifiant", 7)
    idx_creneaux = headers.get("Créneaux restants", 1)
    idx_type     = headers.get("Carte prépayée", 0)
    idx_exp      = headers.get("Valide jusqu'à", 4)

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue
        if len(row) <= max(idx_idd, idx_creneaux): continue
        idd      = row[idx_idd] if len(row) > idx_idd else None
        creneaux = row[idx_creneaux] if len(row) > idx_creneaux else 0
        if idd:
            if idd not in cartes or (creneaux or 0) > cartes[idd].get("creneaux", 0):
                cartes[idd] = {
                    "type":       row[idx_type] if len(row) > idx_type else None,
                    "creneaux":   creneaux or 0,
                    "expiration": row[idx_exp] if len(row) > idx_exp else None,
                }
    return cartes


def _charger_presences(fichier: str) -> dict:
    """Retourne {identifiant: {mois: nb_seances}} (cours coachés uniquement)"""
    wb = openpyxl.load_workbook(fichier)
    ws = wb.active
    presences = defaultdict(lambda: defaultdict(int))

    # Lire les headers pour trouver les bons index
    headers = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if i == 1:
            for j, col in enumerate(row):
                if col:
                    headers[str(col).strip()] = j
            break

    idx_idd      = headers.get("Identifiant", 1)
    idx_activite = headers.get("Activité / Catégorie", 4)
    idx_date     = headers.get("Date", 5)
    idx_presence = headers.get("Présence", 13)

    log.info(f"Présences — colonnes: idd={idx_idd} activite={idx_activite} date={idx_date} presence={idx_presence}")

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue
        if len(row) <= max(idx_idd, idx_activite, idx_date, idx_presence): continue

        idd      = row[idx_idd]
        activite = row[idx_activite] if len(row) > idx_activite else None
        date_str = row[idx_date] if len(row) > idx_date else None
        presence = row[idx_presence] if len(row) > idx_presence else None

        if presence != "Oui" or not idd or not date_str:
            continue
        if activite in ACTIVITES_EXCLUES:
            continue
        if idd in ("nubguest1", "nubguest2", "nubguest3", "nubguest4"):
            continue

        try:
            date = datetime.strptime(str(date_str), "%d-%m-%Y")
            mois = f"{date.year}-{date.month:02d}"
            presences[idd][mois] += 1
        except:
            pass

    return presences


def _derniere_presence(fichier: str, idd: str):
    """Retourne la date de dernière présence d'un membre."""
    wb = openpyxl.load_workbook(fichier)
    ws = wb.active
    derniere = None

    # Lire les headers
    headers = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if i == 1:
            for j, col in enumerate(row):
                if col:
                    headers[str(col).strip()] = j
            break

    idx_idd      = headers.get("Identifiant", 1)
    idx_date     = headers.get("Date", 5)
    idx_presence = headers.get("Présence", 13)

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue
        if len(row) <= max(idx_idd, idx_date, idx_presence): continue
        if row[idx_idd] != idd: continue
        if row[idx_presence] != "Oui": continue
        try:
            date = datetime.strptime(str(row[idx_date]), "%d-%m-%Y")
            if derniere is None or date > derniere:
                derniere = date
        except:
            pass

    return derniere


def _mois_precedent(today: datetime) -> str:
    if today.month == 1:
        return f"{today.year - 1}-12"
    return f"{today.year}-{today.month - 1:02d}"
