"""
Module Resawod — Export automatique via Playwright
Validé le 16/03/2026 — 4/4 exports fonctionnels
"""

import os
import time
import glob
import logging
from playwright.sync_api import sync_playwright

log = logging.getLogger(__name__)

RESAWOD_URL      = os.environ.get("RESAWOD_URL", "https://sport.nubapp.com")
RESAWOD_EMAIL    = os.environ["RESAWOD_EMAIL"]
RESAWOD_PASSWORD = os.environ["RESAWOD_PASSWORD"]
DOWNLOAD_DIR     = os.path.abspath("tmp/downloads")

PAGES = [
    {
        "nom": "utilisateurs",
        "url": f"{RESAWOD_URL}/modules/reports/users/users.php",
        "champs": [
            "Créé le...", "Portable", "Prénom", "Nom", "Email",
            "Date de naissance", "Statut", "Identifiant",
            "Abonnements actifs des utilisateurs", "Tag", "Crédit"
        ]
    },
    {
        "nom": "abonnements",
        "url": f"{RESAWOD_URL}/modules/reports/memberships/users_and_memberships.php",
        "champs": ["Identifiant"]
    },
    {
        "nom": "cartes",
        "url": f"{RESAWOD_URL}/modules/reports/users/users_and_vouchers.php",
        "champs": ["Date d\u0027achat", "Identifiant"]
    },
    {
        "nom": "presences",
        "url": f"{RESAWOD_URL}/modules/reports/activities/activities_and_users.php",
        "champs": ["Identifiant"],
        "filtre_date": True
    },
]


def exporter_resawod() -> dict:
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    for f in glob.glob(f"{DOWNLOAD_DIR}/*.xlsx"):
        os.remove(f)

    fichiers = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--window-size=1920,1080",
                "--force-device-scale-factor=1",
            ]
        )
        # Grande résolution pour que tous les boutons soient visibles
        context = browser.new_context(
            accept_downloads=True,
            viewport={"width": 1920, "height": 1080},
        )
        page = context.new_page()

        log.info("Connexion à Resawod...")
        page.goto(RESAWOD_URL)
        page.wait_for_load_state("networkidle")
        time.sleep(1)
        page.fill('input[name="username"]', RESAWOD_EMAIL)
        page.fill('input[name="password"]', RESAWOD_PASSWORD)
        page.click('button[type="submit"]')
        page.wait_for_load_state("networkidle")
        time.sleep(2)
        log.info(f"Connecté — URL: {page.url}")

        for export in PAGES:
            log.info(f"Export {export['nom']}...")
            try:
                page.goto(export["url"])
                page.wait_for_load_state("networkidle")
                time.sleep(2)
                _fermer_popup(page)
                _supprimer_overlays(page)
                time.sleep(1)
                if export["champs"]:
                    _configurer_champs(page, export["champs"], export["nom"])
                if export.get("filtre_date"):
                    _configurer_filtre_date(page)
                chemin = _exporter_excel(page, export["nom"])
                if chemin:
                    fichiers[export["nom"]] = chemin
                    taille = os.path.getsize(chemin)
                    log.info(f"OK {export['nom']} — {taille} octets")
                    # Logger la structure du fichier pour debug
                    _logger_structure(chemin, export["nom"])
                else:
                    log.error(f"ECHEC {export['nom']}")
            except Exception as e:
                log.error(f"Erreur {export['nom']} : {e}")

        browser.close()

    if len(fichiers) < 4:
        manquants = [p["nom"] for p in PAGES if p["nom"] not in fichiers]
        raise Exception(f"Exports manquants : {manquants}")

    return fichiers


def _logger_structure(chemin: str, nom: str):
    """Log la structure du fichier pour debug."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(chemin)
        ws = wb.active
        nb_lignes = ws.max_row
        nb_cols = ws.max_column
        log.info(f"  Structure {nom}: {nb_lignes} lignes x {nb_cols} colonnes")
        # Logger TOUTES les lignes headers (0,1,2)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 4: break
            log.info(f"  Ligne[{i}]: {[str(c)[:20] if c else '' for c in row]}")
    except Exception as e:
        log.warning(f"  Logger structure {nom}: {e}")


def _fermer_popup(page):
    try:
        page.evaluate("""
            document.querySelectorAll('button').forEach(function(btn) {
                if (btn.textContent.trim() === 'Annuler') btn.click();
            });
        """)
        time.sleep(0.5)
    except: pass


def _supprimer_overlays(page):
    try:
        page.evaluate("document.querySelectorAll('.ui-widget-overlay,.modal-backdrop').forEach(el=>el.remove())")
    except: pass
    time.sleep(0.2)


def _cocher_champ(page, champ):
    champ_js = champ.replace("'", "\\'")
    return page.evaluate(f"""
        (function(){{
            var labels=document.querySelectorAll('label');
            for(var i=0;i<labels.length;i++){{
                if(labels[i].textContent.trim()==='{champ_js}'){{
                    var f=labels[i].getAttribute('for');
                    if(f){{
                        var cb=document.getElementById(f);
                        if(cb&&!cb.checked){{cb.click();return 'coche';}}
                        else if(cb&&cb.checked){{return 'deja_coche';}}
                    }}
                    return 'no_for';
                }}
            }}
            return 'non_trouve';
        }})()
    """)


def _appliquer_filtre_date(page):
    """Applique un filtre date pour les 3 derniers mois."""
    from datetime import datetime, timedelta
    today = datetime.now()
    trois_mois = today - timedelta(days=90)
    date_debut = trois_mois.strftime("%d-%m-%Y")
    try:
        page.evaluate(f"""
            var inputs = document.querySelectorAll('input[type="date"], input.date-field, input[name*="date"]');
            inputs.forEach(function(inp) {{
                if(inp.placeholder && inp.placeholder.includes('début') || inp.name && inp.name.includes('start')) {{
                    inp.value = '{date_debut}';
                    inp.dispatchEvent(new Event('change', {{bubbles: true}}));
                }}
            }});
        """)
        time.sleep(0.5)
    except Exception as e:
        log.debug(f"filtre_date: {e}")


def _configurer_filtre_date(page):
    """Configure le filtre date via page.fill() — début de l'année."""
    from datetime import datetime
    today = datetime.now()
    date_debut = f"01-01-{today.year}"
    date_fin   = today.strftime("%d-%m-%Y")
    log.info(f"  Filtre date : {date_debut} → {date_fin}")
    try:
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1)
        inputs_dates = page.evaluate("""
            (function() {
                var result = [];
                var inputs = document.querySelectorAll('input');
                for(var i=0;i<inputs.length;i++) {
                    var val = (inputs[i].value || '').trim();
                    if(val.match(/^\d{2}-\d{2}-\d{4}$/)) {
                        result.push({index: i, value: val});
                    }
                }
                return result;
            })()
        """)
        log.info(f"  Inputs date trouvés: {inputs_dates}")
        if len(inputs_dates) >= 2:
            idx_debut = inputs_dates[0]["index"]
            idx_fin   = inputs_dates[1]["index"]
            all_inputs = page.locator('input')
            debut_input = all_inputs.nth(idx_debut)
            fin_input   = all_inputs.nth(idx_fin)
            debut_input.scroll_into_view_if_needed()
            debut_input.triple_click()
            debut_input.fill(date_debut)
            debut_input.press("Tab")
            time.sleep(0.3)
            fin_input.triple_click()
            fin_input.fill(date_fin)
            fin_input.press("Enter")
            time.sleep(2)
            page.wait_for_load_state("networkidle")
            time.sleep(2)
            page.evaluate("""
                document.querySelectorAll('button,a').forEach(function(el) {
                    if(el.textContent.trim() === 'Rafraîchir') el.click();
                });
            """)
            time.sleep(3)
            page.wait_for_load_state("networkidle")
            time.sleep(1)
            vals = page.evaluate("""
                (function() {
                    var vals = [];
                    document.querySelectorAll('input').forEach(function(inp) {
                        var v = (inp.value||'').trim();
                        if(v.match(/^\d{2}-\d{2}-\d{4}$/)) vals.push(v);
                    });
                    return vals;
                })()
            """)
            log.info(f"  Dates après filtre: {vals}")
        else:
            log.warning(f"  Pas assez d'inputs date: {len(inputs_dates)}")
    except Exception as e:
        log.warning(f"  Filtre date erreur: {e}")


def _configurer_champs(page, champs, nom):
    log.info(f"  Config champs {nom}...")

    # Utiliser JS pour hover et cliquer sur Champs
    page.evaluate("""
        var btns = document.querySelectorAll('button');
        for(var i=0;i<btns.length;i++){
            if(btns[i].textContent.trim()==='Options'){
                btns[i].dispatchEvent(new MouseEvent('mouseover',{bubbles:true}));
                btns[i].dispatchEvent(new MouseEvent('mouseenter',{bubbles:true}));
                break;
            }
        }
    """)
    time.sleep(0.8)

    page.evaluate("""
        document.querySelectorAll('a,button,li').forEach(function(el){
            if(el.textContent.trim()==='Champs') el.click();
        });
    """)
    time.sleep(1)

    for champ in champs:
        _cocher_champ(page, champ)
        time.sleep(0.1)

    page.evaluate("""
        var btns=document.querySelectorAll('button');
        for(var i=0;i<btns.length;i++){
            if(btns[i].textContent.trim()==='Accepter'&&btns[i].offsetParent!==null){
                btns[i].click();break;
            }
        }
    """)
    time.sleep(1)


def _exporter_excel(page, nom):
    # Utiliser JS pour déclencher le hover sur Options
    page.evaluate("""
        var btns = document.querySelectorAll('button');
        for(var i=0;i<btns.length;i++){
            if(btns[i].textContent.trim()==='Options'){
                btns[i].dispatchEvent(new MouseEvent('mouseover',{bubbles:true}));
                btns[i].dispatchEvent(new MouseEvent('mouseenter',{bubbles:true}));
                break;
            }
        }
    """)
    time.sleep(0.8)

    # Cliquer sur Exporter en Excel via JS
    try:
        with page.expect_download(timeout=60000) as dl:
            page.evaluate("""
                document.querySelectorAll('a,button,li').forEach(function(el){
                    if(el.textContent.trim()==='Exporter en Excel') el.click();
                });
            """)
        download = dl.value
        chemin = os.path.join(DOWNLOAD_DIR, f"{nom}.xlsx")
        download.save_as(chemin)
        return chemin
    except Exception as e:
        log.error(f"Erreur téléchargement {nom}: {e}")
        return None
